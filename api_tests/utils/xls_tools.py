import os.path

import openpyxl


class XlsTools(object):
    def __init__(self, file_path, title=None):
        self.file_path = file_path
        try:
            self.workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(file_path)
            self.workbook = openpyxl.load_workbook(file_path)

        self.sheet = self.workbook.active
        if title:
            self.write_row(title, row=1)
            self.save()

    def read_cell(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def write_cell(self, row, column, value):
        self.sheet.cell(row=row, column=column).value = value

    def write_row(self, data, row=0):
        new_row = row
        if new_row == 0:
            last_row = self.sheet.max_row
            new_row = last_row + 1
        for idx, value in enumerate(data, start=1):
            self.write_cell(new_row, idx, value)

    def delete_row(self, row):
        self.sheet.delete_rows(row)

    def read_all_rows(self):
        all_rows = []
        for row in self.sheet.iter_rows(values_only=True):
            all_rows.append(row)
        return all_rows

    def save(self):
        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()


if __name__ == '__main__':
    xls = XlsTools("../data/data.xlsx")
    res = xls.read_all_rows()
    for i in res:
        print(i)
